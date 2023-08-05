# Imports from python.
import csv
import os
import subprocess


# Imports from Django.
from django.core.management.base import BaseCommand


# Imports from other dependencies.
from election.models import Candidate
from election.models import CandidateElection
from election.models import Election
from election.models import Race
from entity.models import Person
from government.models import Party
from openpyxl import load_workbook
from time import sleep


class Command(BaseCommand):
    help = "Bootstrap development."

    def write_csvs(self):
        wb_path = os.path.join(self.cmd_path, "../../bin/candidates.xlsx")
        wb = load_workbook(wb_path)

        for name in wb.sheetnames:
            csv_path = os.path.join(
                self.cmd_path,
                "../../bin/csv",
                "{}-candidates.csv".format(name),
            )

            with open(csv_path, "w") as writefile:
                subprocess.Popen(
                    ["in2csv", wb_path, "-f", "xlsx", "--sheet", name],
                    stdout=writefile,
                )

    def get_candidates(self):
        sleep(2)
        with open(
            os.path.join(self.cmd_path, "../../bin/csv/house-candidates.csv")
        ) as f:
            reader = csv.DictReader(f)

            for row in reader:
                race = self.get_house_race(row)
                self.process_row(row, race)

        with open(
            os.path.join(self.cmd_path, "../../bin/csv/senate-candidates.csv")
        ) as f:
            reader = csv.DictReader(f)

            for row in reader:
                race = self.get_senate_race(row)
                self.process_row(row, race)

        with open(
            os.path.join(
                self.cmd_path, "../../bin/csv/governor-candidates.csv"
            )
        ) as f:
            reader = csv.DictReader(f)

            for row in reader:
                race = self.get_governor_race(row)
                self.process_row(row, race)

    def get_house_race(self, row):
        print("house", row["State"], row["District"])
        district_int = float(row["District"])
        if district_int < 10:
            district_str = row["District"].zfill(2)
        else:
            district_str = row["District"]

        race = Race.objects.get(
            cycle__slug="2018",
            office__division__parent__label=row["State"],
            office__division__code=district_str,
            office__body__slug="house",
            special=False,
        )

        return race

    def get_senate_race(self, row):
        print("senate", row["State"], row["Special"])

        kwargs = {
            "cycle__slug": "2018",
            "office__division__label": row["State"],
            "office__body__slug": "senate",
        }

        if row["Special"] == "True":
            kwargs["special"] = True
        else:
            kwargs["special"] = False

        race = Race.objects.get(**kwargs)

        return race

    def get_governor_race(self, row):
        print("governor", row["State"])
        race = Race.objects.get(
            cycle__slug="2018",
            office__division__label=row["State"],
            office__body=None,
            special=False,
        )

        return race

    def process_row(self, row, race):
        person, created = Person.objects.get_or_create(
            first_name=row["First"], last_name=row["Last"]
        )

        party, created = Party.objects.get_or_create(label=row["Party"])

        candidate, created = Candidate.objects.get_or_create(
            person=person, race=race, party=party, incumbent=row["Incumbent"]
        )

        election = Election.objects.get(
            race=race, election_day__date="2018-11-06"
        )

        CandidateElection.objects.get_or_create(
            candidate=candidate, election=election, aggregable=False
        )

    def handle(self, *args, **options):
        self.cmd_path = os.path.dirname(os.path.realpath(__file__))

        self.write_csvs()
        self.get_candidates()
