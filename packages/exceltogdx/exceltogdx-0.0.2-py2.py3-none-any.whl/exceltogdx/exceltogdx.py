import gdxpds
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from io import BytesIO
import os
import re


def xlsdynamicecke(typ, cell, rdim, cdim, sheetname, wb):
    '''
    Returns a list of row and col of bottom-left corner of a table in pandas indexing format (from zero to inf).
    It stops when there is an empty cell in index (rows) or headings (columns).
    typ: string 'set' or 'par'
    cell: string in excel format of top-right table corner cell.
    rdim: indicates the number of columns from the beginning are sets
    cdim: indicates the number of rows from the top are sets
    sheetname: self-explanatory
    wb: is the workbook of an excel file instance of 'from openpyxl import load_workbook'
    eg. xlsdynamicecke('set', C5', 1, 0, 'sheet1', workbook.object)
    return set or table coord.
    '''
    cell = cell.upper()
    sheet = wb[sheetname]
    string = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'

    def col2num(letters):
        '''
        column letter to column number
        '''
        num = 0
        for c in letters:
            if c in string:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def colnum_string(n):
        strings = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            strings = chr(65 + remainder) + strings
        return strings

    def natural_keys(text):
        '''
        alist.sort(key=natural_keys) sorts in human order
        http://nedbatchelder.com/blog/200712/human_sorting.html
        (See Toothy's implementation in the comments)
        '''
        def atoi(text):
            return int(text) if text.isdigit() else text
        return [atoi(c) for c in re.split(r'(\d+)', text)]

    cut = 0
    for s in cell:
        if s in string:
            cut += 1
        else:
            break
    rowstr = cell[cut:]
    colstr = cell[:cut]

    row = int(rowstr)
    col = col2num(colstr.upper())
    if typ == 'par':
        if cdim == 0:
            j = 0
            for i, r in enumerate(sheet.iter_rows(min_row=row, min_col=col, max_col=col, values_only=True)):
                j = i
                if r[0] is None:
                    j = i - 1
                    break
            max_col = rdim + 1
            max_row = row + j
            rng = colnum_string(col) + str(row - 1) + ':' + colnum_string(max_col) + str(max_row)
            data = sheet[rng]
            output = [[cells.value for cells in row] for row in data]
            print(rng)
            print(output[:3])
        else:
            j = 0
            for i, c in enumerate(sheet.iter_cols(min_row=row, max_row=row, min_col=col+rdim, values_only=True)):
                j = i
                if c[0] is None:
                    j = i - 1
                    break
            max_col = col + j + rdim
            for i, r in enumerate(sheet.iter_rows(min_row=row+cdim+1, min_col=col, max_col=col, values_only=True)):
                j = i
                if r[0] is None:
                    j = i - 1
                    break
            max_row = row + j + cdim + 1
            rng = cell + ':' + colnum_string(max_col) + str(max_row)
            data = sheet[rng]
            output = [[cells.value for cells in row] for row in data]
            print(rng)
            print(output[:3])
    elif typ == 'set':
        setls = []
        if rdim == 1:
            for i, r in enumerate(sheet.iter_rows(min_row=row, min_col=col, max_col=col, values_only=True)):
                if r[0] is not None:
                    setls.append(r[0])
                else:
                    break
            if all([isinstance(s, (int, float)) for s in list(set(setls))]):
                output = sorted(list(set(setls)))
            else:
                output = sorted(list(set(setls)), key=natural_keys)

        elif cdim == 1:
            for i, c in enumerate(sheet.iter_cols(min_row=row, max_row=row, min_col=col, values_only=True)):
                if c[0] is not None:
                    setls.append(c[0])
                else:
                    break
            if all([isinstance(s, (int, float)) for s in list(set(setls))]):
                output = sorted(list(set(setls)))
            else:
                output = sorted(list(set(setls)), key=natural_keys)
        else:
            raise ValueError('Set must have either rdim or cdim as 1, check dim in py sheet')
    del sheet
    return output


def exceltogdx(excel_file, gdx_file, csv_file=None, csv_copy=None):
    '''
    excel_file: input file path
    gdx_file: output file path
    csv_file: if None, it looks at excel file to find sheet with name 'py'
              that contains the instructions to get sets and parameters.
              Otherwise, csv file path.
    csv_copy: indicate folder where csv files are saved. None (Default): no csv files are created.
    '''
    if csv_file is None:
        mapping = pd.read_excel(excel_file, sheet_name='py', index_col='symbol')
    else:
        mapping = pd.read_csv(csv_file, index_col='symbol')

    print("loading excel file...")
    with open(excel_file, 'rb') as f:
        datas = BytesIO(f.read())
    wb = load_workbook(datas, data_only=True)
    dc = {}
    df = pd.DataFrame()
    for k, v in mapping.iterrows():
        print(v['type'],': ', k)
        xlsvalues = xlsdynamicecke(v['type'], v['startcell'], v['rdim'], v['cdim'], v['sheet_name'], wb)
        if v['type'] == 'par':
            df = pd.DataFrame(xlsvalues)
            if v['cdim'] == 0:
                df = df.T.set_index(0, append=False).T
                try:
                    df = df.set_index(df.columns[list(range(v['rdim']))].to_list())
                except KeyError:
                    raise KeyError("each rdim in parameter '{}' must have a heading (Don't leave it empty), not required for cdim".format(k))
                df.index.names = list(range(1,df.index.nlevels+1))
            elif v['cdim'] == 1:
                df = df.T.set_index(0, append=False).T
                try:
                    df = df.set_index(df.columns[list(range(v['rdim']))].to_list())
                except KeyError:
                    raise KeyError("each rdim in parameter '{}' must have a heading (Don't leave it empty), not required for cdim".format(k))
                df = df.stack([0]*df.columns.nlevels)
                df.index.names = list(range(1,df.index.nlevels+1))
                df = pd.DataFrame(df)
            elif v['cdim'] > 1:
                df = df.T.set_index(list(range(v['cdim'])), append=False).T
                try:
                    df = df.set_index(df.columns[list(range(v['rdim']))].to_list())
                except KeyError:
                    raise KeyError("each rdim in parameter '{}' must have a heading (Don't leave it empty), not required for cdim".format(k))
                df = df.stack([0]*df.columns.nlevels)
                df.index.names = list(range(1,df.index.nlevels+1))
                df = pd.DataFrame(df)
            else:
                raise Exception('is "{}" a parameter?, verify cdim on "py" sheet. cdim must be positive integer'.format(k))
            df = df.reset_index().rename(columns={df.columns.to_list()[-1]: 'value'})
            df.loc[df[df['value'] == 'inf'].index, 'value'] = np.inf
            df.loc[:,[c for c in df.columns if (c != 'value' and df[c].dtypes == float)]] = df[[c for c in df.columns if (c != 'value' and df[c].dtypes == float)]].astype(int)
            df.loc[:,[c for c in df.columns if c != 'value']] = df[[c for c in df.columns if c != 'value']].astype(str)
            dc[k] = df.rename(columns={c: '*' for c in df.columns if c != 'value'})
        elif v['type'] == 'set':
            df = pd.DataFrame({'*': xlsvalues})
            df.loc[:, 'value'] = 'True'
            df.dropna(inplace=True)
            dc[k] = df
        if csv_copy is not None:
            os.makedirs(csv_copy, exist_ok=True)
            name = v['type'] + '_' + k + '.csv'
            df.to_csv(os.path.join(csv_copy, name), index=False)
    os.makedirs(os.path.abspath(os.path.join(gdx_file, os.pardir)), exist_ok=True)
    print('generating gdx file...')
    gdxpds.to_gdx(dc, gdx_file)
    print('Done!')
    return dc