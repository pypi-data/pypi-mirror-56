"""This module focuses on getting and validating the path for the RTM
worksheet."""

# --- Standard Library Imports ------------------------------------------------
import datetime
import tkinter as tk
from pathlib import Path
from tkinter import filedialog
import os

# --- Third Party Imports -----------------------------------------------------
import click
import openpyxl
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.comments import Comment

# --- Intra-Package Imports ---------------------------------------------------
from rtm.containers.markup import CellMarkup
from rtm.main import exceptions as exc
from rtm.main.versions import get_version_check_message


def get_rtm_path(path_option='default') -> Path:
    """Prompt user for RTM workbook location. Return path object."""
    if path_option == 'default':
        path = get_new_path_from_dialog()
        required_extensions = '.xlsx .xls'.split()
        if str(path) == '.':
            raise exc.RTMValidatorFileError("\nError: You didn't select a file")
        if path.suffix not in required_extensions:
            raise exc.RTMValidatorFileError(
                f"\nError: You didn't select a file with "
                f"a proper extension: {required_extensions}"
            )
        click.echo(f"\nThe RTM you selected is {path}")
        return path
    elif isinstance(path_option, Path):
        return path_option


def get_new_path_from_dialog() -> Path:
    """Provide user with dialog box so they can select the RTM Workbook"""
    root = tk.Tk()
    root.withdraw()
    path = Path(filedialog.askopenfilename())
    return path


def get_workbook(path):
    return openpyxl.load_workbook(filename=str(path), data_only=True)


def get_worksheet(workbook, worksheet_name):
    ws = None
    for sheetname in workbook.sheetnames:
        if sheetname.lower() == worksheet_name.lower():
            ws = workbook[sheetname]
    if ws is None:
        raise exc.RTMValidatorFileError(
            f"\nError: Workbook does not contain a '{worksheet_name}' worksheet"
        )
    return ws


def now_str(pretty=False):
    if pretty:
        return datetime.datetime.now().strftime("%d %B %Y, %I:%M %p")
    else:
        return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


def get_save_path(original_path, modify_original_file=False):
    """Get the full file path. Create subdirectory if necessary."""

    # --- When modifying original file ----------------------------------------
    if modify_original_file:
        return original_path

    # --- When saving a copy of the original file -----------------------------
    original_path = Path(original_path)
    original_directory = original_path.parent
    subdirectory = original_directory/'rtm_validator_results'
    subdirectory.mkdir(exist_ok=True)
    file_name = f'{now_str()}_{original_path.name}'
    return subdirectory / file_name


def get_cell_comment_string(comments):
    # comments is a list of title/explanation pairs
    titles_and_comments = [f"{comment[0].upper()}\n{comment[1]}" for comment in comments]
    comments_string = '\n\n'.join(titles_and_comments)
    return f"{now_str(pretty=True)}\n\n{comments_string}"


def mark_up_excel(path, wb, ws_procedure, markup_content: dict, modify_original_file=False):
    # Comments fall in two categories:
    #   PROCEDURE BASED REQUIREMENTS: These are comments bound to record in a specific field.
    #       The cell gets highlighted orange and a comment explains the error.
    #   README: These are not directed at a specific cell.
    #       These will generate new rows inserted at the top of the worksheet.

    bg_error = Color('00edc953')  # yellow-ish background color for procedure errors
    fg_good = Color('0025c254')  # green-ish text color for readme comments
    fg_error = Color('00bf2f24')  # red-ish text color for readme errors

    # --- Procedure markup ----------------------------------------------------
    for location, comments in markup_content.items():
        if not isinstance(location, str):  # i.e. if this items has a row/col location
            cell = ws_procedure.cell(*location)
            # cell.style = style_error
            cell.fill = PatternFill(patternType='solid', fgColor=bg_error)
            cell.comment = Comment(get_cell_comment_string(comments), "RTM Validator")

    # --- Set up README errors ------------------------------------------------
    general_errors = []
    for field_name, comments in markup_content.items():
        if isinstance(field_name, str):
            general_errors.append(CellMarkup(field_name.upper(), is_error=True))
            for comment in comments:
                comment_str = f"{comment[0].upper()}: {comment[1]}"
                general_errors.append(CellMarkup(comment_str, is_error=True, indent=True))

    # --- Set up README comments ----------------------------------------------
    readme_text = [
        CellMarkup("RTM VALIDATOR", size=24),
        CellMarkup(f"{now_str(pretty=True)}"),
        CellMarkup("All images and attachments have been removed from this workbook."),
        CellMarkup(),
        CellMarkup("Cells highlighted orange require attention."),
        CellMarkup("See the cell's note/comment for details."),
        CellMarkup(),
        CellMarkup("To improve readability, convert notes to comments:"),
        CellMarkup("Go to the Review tab", indent=True),
        CellMarkup("Click on Notes, select Convert to Comments", indent=True),
        CellMarkup(),
    ] + get_version_check_message()
    if general_errors:
        readme_text += [
                              CellMarkup(),
                              CellMarkup("General Errors:"),
                              CellMarkup(),
                          ] + general_errors

    # --- create and write to README sheet ------------------------------------
    readme = 'README'
    ws_readme = wb.create_sheet(readme, 0)
    for row, comment in enumerate(readme_text, 1):
        cell = ws_readme.cell(row, 1, comment.comment)
        cell.alignment = Alignment(
            wrapText=False,
            indent=3 if comment.indent else 0,
        )
        cell.font = Font(
            color=fg_error if comment.is_error else fg_good,
            size=comment.size,
            bold=True,
        )
        if comment.size:
            ws_readme.row_dimensions[row].height = comment.size * 1.4

    # --- Delete Unmarked Sheets ----------------------------------------------
    if not modify_original_file:
        for worksheet in wb.worksheets:
            if worksheet not in [ws_procedure, ws_readme]:
                wb.remove(worksheet)
    sheet_index = 0
    wb.active = sheet_index

    # --- Save ----------------------------------------------------------------
    save_path = get_save_path(path, modify_original_file)
    wb.save(save_path)
    os.startfile(save_path)
    # open(save_path)


def row_heights(ws):
    heights = [ws.row_dimensions[index+1].height for index in range(ws.max_row)]
    return [15 if height is None else height for height in heights]


if __name__ == '__main__':
    pass
