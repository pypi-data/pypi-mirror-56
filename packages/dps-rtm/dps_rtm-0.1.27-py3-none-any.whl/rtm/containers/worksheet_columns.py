"""This module defines the worksheet column class and the custom sequence
containing all the worksheet columns. Worksheet columns are containers housing
a single column from a worksheet."""

# --- Standard Library Imports ------------------------------------------------
from collections import namedtuple
from typing import List

# --- Third Party Imports -----------------------------------------------------
import openpyxl

# --- Intra-Package Imports ---------------------------------------------------
import rtm.main.config as config


WorksheetColumn = namedtuple("WorksheetColumn", "header values position column")
# header: row 1
# values: list of cell values starting at row 2
# position: similar to column number, but starts as zero, like an index
# column: column number (start at 1)


class WorksheetColumns:

    def __init__(self, worksheet):

        # --- Attributes ------------------------------------------------------
        self.max_row = worksheet.max_row
        self.height = self.max_row - config.header_row
        self.cols = []

        # --- Convert Worksheet to WorksheetColumn objects ----------------
        start_column_num = 1
        for position, col in enumerate(range(start_column_num, worksheet.max_column + 1)):
            column_header = str(worksheet.cell(config.header_row, col).value)
            column_values = tuple(str(worksheet.cell(row, col).value) for row in range(config.header_row+1, self.max_row + 1))
            ws_column = WorksheetColumn(
                header=column_header, values=column_values, position=position, column=col
            )
            self.cols.append(ws_column)

    def get_first(self, header_name):
        """returns the first worksheet_column that matches the header"""
        matches = get_matching_worksheet_columns(self, header_name)
        if len(matches) > 0:
            return matches[0]
        else:
            return None

    # --- Sequence ------------------------------------------------------------
    def __getitem__(self, index):
        return self.cols[index]

    def __len__(self):
        return len(self.cols)


def get_matching_worksheet_columns(sequence_worksheet_columns, field_name) -> List[WorksheetColumn]:
    """Called by constructor to get matching WorksheetColumn objects"""
    matching_worksheet_columns = [
        ws_col
        for ws_col in sequence_worksheet_columns
        if ws_col.header.lower() == field_name.lower()
    ]
    return matching_worksheet_columns
