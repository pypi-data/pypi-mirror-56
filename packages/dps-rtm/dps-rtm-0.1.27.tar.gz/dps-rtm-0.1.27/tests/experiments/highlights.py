"""Here I experiment with how to add styles/colors/etc to excel cells."""

# See https://openpyxl.readthedocs.io/en/stable/styles.html


from pathlib import Path
import openpyxl
from openpyxl.comments import Comment
import os
import random
import datetime


# --- get path ----------------------------------------------------------------
path = Path(__file__).parent.parent / "test_rtm.xlsx"

# --- get workbook ------------------------------------------------------------
wb = openpyxl.load_workbook(str(path))

# --- get worksheet -----------------------------------------------------------
ws_name = 'colors'
try:
    ws = wb[ws_name]
except KeyError:
    ws = wb.create_sheet(ws_name)

# --- write to rand cell ------------------------------------------------------
row = random.randrange(1,6)
col = random.randrange(1,6)
value = f'({row}, {col})'
cell = ws.cell(row, col, value)

# --- modify cell style -------------------------------------------------------
# styles = [f'Accent{num}' for num in range(1,7)]
# style = random.choice(styles)
style = 'Accent2'
cell.style = style

# --- add cell comment --------------------------------------------------------
comment_value = f'You already know this, but this is cell {value}'
comment_author = f'RTM Validator'  # , {datetime.datetime.now()}'
cell.comment = Comment(comment_value, comment_author)

# --- save --------------------------------------------------------------------
wb.save(path)

# --- modify comment ----------------------------------------------------------
wb = openpyxl.load_workbook(str(path))
ws = wb[ws_name]
cell = ws.cell(row, col, value)

comment_old = cell.comment
if comment_old is None:
    comment_value_new = ''
else:
    comment_value_new = comment_old.text + '\n\n'
comment_value_new += 'This is a second line.'
comment_new = Comment(
    text=comment_value_new,
    author=comment_author
)
cell.comment = comment_new
wb.save(path)

# --- open file for inspection ------------------------------------------------
os.startfile(path)
