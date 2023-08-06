"""Read confusion-count data stored in Excel 2007- (xlsx) workbook file.

Each Excel workbook file may include stimulus-response data for
one or more subject(s), tested in one or more test-factor combinations.

Each workbook sheet should include stimulus-response data for
exactly ONE subject in ONE combination of test-factor categories,
for ONE list of stimuli
and ONE corresponding list of responses.

Data elements must be found in the SAME positions in all sheets.

If any cells are empty in the designated stimulus range,
the sheet is assumed empty and simply discarded.

Response data may be saved either in the form of a response-count matrix,
or simply as a list including one response label for each stimulus label.
The order of stimuli and responses are arbitrary, and may differ
across sheets and workbook files.

A response-count matrix may be stored either with
one ROW for each stimulus category, OR one COLUMN for each stimulus,
as defined by the stim_range and resp_range parameters.


*** Usage Example:

cm_file = StimRespFile(file_path,
        sheets=[f'Subject{i} for i in range(10)],
        subject='sheet',
        stim_range='C4:C20',
        resp_range='D3:U3',
        count_range='D4:U20',
        test_factors=dict(HA=['A', 'B'], SNR=['low', 'high']),
        HA='A2')

The StimRespFile properties define sheet locations for
subject, stimulus labels, response labels, response counts,
and test-factor category for each test-factor.

The location of single data elements can be either a cell address
or 'sheet', indicating that the sheet name is to be used.

The parameter sheets is a list of workbook sheets to be searched for data.

In this example, subject='sheet' indicates that the sheet name is interpreted as subject id.
The parameter stim_range is a column containing stimulus category labels,
and resp_range is a heading row with response labels.

Each cell in count_range contains an integer response count
for the stimulus in the corresponding position, such that
count[i, j] is the response count for the response label in the j-th cell in resp_range,
for the stimulus label in the i-th cell in stim_range.
Thus, the shapes of stim_range, resp_range, and count_range must match each other.

The parameter test_factors includes possible categories for each given test factor.

In this example, cell 'A2' specifies a category for test-factor 'HA'.
The file path string will be examined to find
one of the allowed categories for the remaining test-factor 'SNR',
which is not assigned any cell address.


*** Version History:

2018-09-14, first functional version
"""

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.exceptions import InvalidFileException

from collections import Counter

from . import cm_base


class FileFormatError(cm_base.FileReadError):
    """Format error causing non-usable data"""


class ParameterError(cm_base.FileReadError):
    """Error in calling parameters causing non-usable data"""


class StimRespFile(cm_base.StimRespFile):
    """Interface to Excel xlsx workbook file storing stimulus-response data.
    Each sheet in the workbook must include stimulus-response data for
    ONE subject in ONE test condition.
    All fields MUST be stored at the same cell addresses in all sheets.
    """
    def __init__(self, file_path, stim_range, resp_range, count_range=None,
                sheets=None, subject='sheet', test_factors=None, **tf_address):
        """File interface to data read from an excel file,
        in format compatible with class StimRespRecord.
        :param file_path: Path to file for reading
        :param stim_range: single COLUMN or ROW cell range, like 'A3:A50' or 'A3:Z3',
            with stim_range label strings. The range may include EITHER
            a heading row or column for a confusion matrix, OR
            a list with labels of presented phonemes, in presentation order.
        :param resp_range: single ROW or COLUMN cell range, with response labels, EITHER
            with unique labels as column or row headings for a confusion matrix, OR
            a list parallel to stim_range, with one response label for each presentation
        :param count_range: (optional) range of response counts in a confusion matrix
        :param sheets: (optional) list of sheet names to be searched for data.
        :param subject: (optional) location of subject identification string,
            = either a cell address like 'A1', OR
            = 'sheet', indicating that the sheet name is used as subject code.
        :param test_factors: (optional) dict with elements (test_factor, tf_cat_list),
            as saved in a ConfMatrixFrame instance,
            used only if some test_factor is not defined in a sheet cell
        :param tf_address: (optional) dict with elements (tf, location), where
            tf is a string identifying one test factor,
            location is a string with a cell address like 'D4',
            OR 'sheet', indicating that the sheet name is the test-factor category.
            For additional test_factors.keys(), if not yet found,
            search for the category as a sub-string in file_path.
        :return: generator yielding rec = dict with fields
            subject, test_cond, stim_resp_count
        """
        super().__init__(file_path, test_factors)
        _check_sr_range(stim_range, resp_range, count_range)
        _check_test_cond(tf_address)
        # raising ArgumentError if any problem found
        self.stim_range = stim_range
        self.resp_range = resp_range
        self.count_range = count_range
        self.sheets = sheets
        self.subject = subject
        self.tf_address = tf_address

    def __iter__(self):
        """Generator yielding data from an excel file,
        in dicts compatible with class StimRespRecord.
        :return: generator yielding rec = dict with fields
            subject, test_cond, stim_resp_count
        """
        file_string = str(self.file_path)
        try:
            wb = load_workbook(file_string, read_only=True)
        except InvalidFileException:
            raise FileFormatError(f'Cannot load workbook from file {self.file_path.name}')
        if self.sheets is None:
            sheets = wb.sheetnames
        else:
            sheets = set(self.sheets) & set(wb.sheetnames)
        if len(sheets) == 0:
            raise FileFormatError('Sheets not found in {file_path.stem}')
        path_test_cond = self.path_test_cond()
        for sheet_name in sheets:
            ws = wb[sheet_name]
            sr = _get_sr_dict(ws, self.stim_range, self.resp_range, self.count_range)
            subject_id = _get_field(ws, self.subject)
            sheet_test_cond = _get_test_cond(ws, self.tf_address)
            tc = path_test_cond.copy()
            tc.update(sheet_test_cond)
            if len(sr) > 0 and all(s is not None for s in sr.keys()):
                res = {'subject': subject_id,
                       'test_cond': tc,
                       'stim_resp_count': sr}
                yield res


# --------------------------------------------------- help sub-functions

def _cell_range_shape(range):
    cs = CellRange(range).size
    return cs['rows'], cs['columns']


def _check_sr_range(stim_range, resp_range, count_range):
    stim_shape = _cell_range_shape(stim_range)
    resp_shape = _cell_range_shape(resp_range)
    if count_range is None:
        if stim_shape != resp_shape:
            raise ParameterError('stim and resp must have equal size')
    else:
        count_shape = _cell_range_shape(count_range)
        if stim_shape[0] == 1 and resp_shape[1] == 1:
            if stim_shape[1] != count_shape[1] or resp_shape[0] != count_shape[0]:
                raise ParameterError('stim and resp ranges must match count range')
        elif stim_shape[1] == 1 and resp_shape[0] == 1:
            if stim_shape[0] != count_shape[0] or resp_shape[1] != count_shape[1]:
                raise ParameterError('stim and resp ranges must match count range')
        else:
            raise ParameterError('stim and resp ranges must be single row or column')


def _check_test_cond(tf_address):
    for (tf, addr) in tf_address.items():
        if addr != 'sheet' and _cell_range_shape(addr) != (1, 1):
            raise ParameterError(f'Test_factor {tf} must be a single cell address')


def _get_field(ws, cell):
    """Get contents in ONE cell or in sheet title
    :param ws: a worksheet
    :param cell: one cell address or 'sheet'
    :return: cell contents
    """
    if cell == 'sheet':
        return ws.title
    c = ws[cell]
    if type(c) is tuple:
        raise FileFormatError('Multiple cells where single cell is required')
    else:
        return c.value


def _get_test_cond(ws, tf_address):
    """Get tuple of test conditions
    :param ws: a worksheet
    :param tf_address: dict with elements (tf, address), where
        tf is one test-factor label
        address is a cell address or 'sheet' or 'path'
    :return: tuple of (tf, tf_category) pairs
    """
    def gen_tc():
        "generator of test-condition tuples"
        for (tf, addr) in tf_address.items():
            tc = _get_field(ws, addr)
            if tc is not None:
                yield (tf, tc)
    # -----------------------------------------------------------------

    return dict(tf_c for tf_c in gen_tc())


def _get_sr_dict(ws, s_range, r_range, c_range):
    """Get stim-response dict from give work-sheet
    :param ws: a worksheet
    :param s_range:
    :param r_range:
    :param c_range:
    :return: sr = dict with elements (s: r_dict), where
        s is a stim label,
        r_dict is a dict with elements (r, count)
    """
    stim = np.array(ws[s_range])
    resp = np.array(ws[r_range])
    # = arrays of cell objects
    stim_is_row = stim.shape[0] == 1
    stim = [s.value for s in stim.flatten()]
    resp = [r.value for r in resp.flatten()]
    if c_range is None:
        sr = {s: Counter() for s in set(stim)}
        for (s, r) in zip(stim, resp):
            sr[s].update({r: 1})
    else:
        c = np.array([[c_ij.value for c_ij in c_i]
                      for c_i in ws[c_range]])
        if stim_is_row:
            c = c.T
        sr = {s: {r: c_sr for (r, c_sr) in zip(resp, c_s)
                  if c_sr is not None and c_sr > 0}
              for (s, c_s) in zip(stim, c)}
    return sr

