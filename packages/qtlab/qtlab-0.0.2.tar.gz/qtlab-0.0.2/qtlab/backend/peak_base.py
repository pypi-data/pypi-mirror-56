#!/usr/bin/python3
import os
import math
import pandas

from openpyxl import Workbook
from openpyxl.chart import Reference, Series, ScatterChart
from openpyxl.drawing.colors import ColorChoice
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class Chart(ScatterChart):
    def __init__(self, parent, x_min=None, x_max=0, peaks=True):
        super().__init__()
        self.parent = parent
        self.labels = parent.labels
        self.ws = parent.ws
        self.__elements__ = ('scatterStyle', 'varyColors', 'ser', 'dLbls', 'axId',)
        self._setup_axis(x_min, x_max)

        if peaks:
            self.legend = None
            self.x_axis.majorGridlines = None
            self.ws_peaks = parent.create_sheet("__peaks__")
            self.ws_peaks.sheet_state = 'hidden'
        self._add_series(peaks)

    def _add_peaks(self, column):
        x_col = get_column_letter(column)
        y_col = get_column_letter(column+1)
        index, peaks = self._find_peaks(y_col)

        for i, peak in enum(index, start=0, step=2):
            x_peak = int(self.ws[f"{x_col}{peak+2}"].value)
            self.ws_peaks[f"{x_col}{i+1}"] = x_peak
            self.ws_peaks[f"{x_col}{i+2}"] = x_peak
            self.ws_peaks[f"{y_col}{i+1}"] = 0
            self.ws_peaks[f"{y_col}{i+2}"] = self.y_axis.scaling.max

            x_values = Reference(self.ws_peaks, min_col=column, min_row=i+1, max_row=i+2)
            y_values = Reference(self.ws_peaks, min_col=column+1, min_row=i+1, max_row=i+2)
            color = self.series[-1].graphicalProperties.line.solidFill.srgbClr

            series = Series(y_values, x_values)
            series.graphicalProperties.line.width = 12721*1  # Conversion of pts to EMUs
            series.graphicalProperties.line.prstDash = "sysDash"
            series.graphicalProperties.line.solidFill = ColorChoice(srgbClr=color)
            series.smooth = True
            self.series.append(series)

    def _add_series(self, peaks):
        self.colors = ["386192", "953937", "769140", "614a7d", "35849a", "c17230",
                        "4575af", "ad3b38", "8dac4c", "745995", "3f9eb7", "e58536"]

        for i in range(1, self.ws.max_column, 2):
            x_values = Reference(self.ws, min_col=i, min_row=2, max_row=self.ws.max_row)
            y_values = Reference(self.ws, min_col=i+1, min_row=1, max_row=self.ws.max_row)
            series = Series(y_values, x_values, title_from_data=True)
            series.graphicalProperties.line.width = 12721*2  # Conversion of pts to EMUs
            series.graphicalProperties.line.solidFill = ColorChoice(srgbClr=self.colors[0])
            series.smooth = True
            self.series.append(series)
            if peaks:
                self._add_peaks(i)
            self.colors.append(self.colors.pop(0))

    def _find_peaks(self, y_col):
        values = []
        for cell in self.ws[y_col]:
            if isinstance(cell.value, float):
                if not pandas.isnull(cell.value):
                    values.append(cell.value)
        values = pandas.array(values)
        index, peaks = self._filter_peaks(values)
        return index, peaks

    def _setup_axis(self, x_min=None, x_max=0, y_min=0, y_max=0):
        scale = self.Scale(self, x_min, x_max, y_min, y_max)
        self.y_axis.scaling.min = scale.y_min
        self.x_axis.scaling.min = scale.x_min
        self.y_axis.scaling.max = scale.y_max
        self.x_axis.scaling.max = scale.x_max
        self.x_axis.number_format = "0"
        self.y_axis.number_format = "0"
        if scale.y_max < 15:
            self.y_axis.majorUnit = 1

    class Scale:
        def __init__(self, parent, x_min, x_max, y_min, y_max):
            self.x_min, self.x_max, self.y_min, self.y_max = x_min, x_max, y_min, y_max
            self.ws = parent.ws

            baseSheet = parent.parent
            x_values, minimums, maximums = baseSheet.mins_maxs(1, self.ws.max_column)
            self.x_max = max(maximums) if not x_max else x_max
            self.x_min = min(minimums) if x_min is None else x_min
            if not y_max:
                self.y_max = self._y_max(x_values, x_min, x_max)

        def _rows_limits(self, x_values, x_start, x_stop):
            # Reverse start/stop mark for descending x values
            for col in x_values:
                first = x_values[col][0]
                last = x_values[col][-1]
                break
            if first > last:
                x_stop, x_start = x_start, x_stop

            rows_limits = {}
            for i in range(1, self.ws.max_column, 2):
                letter = get_column_letter(i)
                start, stop = None, None
                for cell in self.ws[letter]:
                    if start and stop: break
                    begin = x_start if x_start else min(x_values[letter])
                    end = x_stop if x_stop else max(x_values[letter])
                    begin, end = int(begin), int(end)
                    try:
                        value = int(cell.value)
                        if value == begin:
                            start = cell.row
                        elif value == end:
                            stop = cell.row
                    except ValueError:
                        pass

                start = 2 if start is None else start
                stop = cell.row if stop is None else stop
                if start > stop:
                    start, stop = stop, start
                y_letter = get_column_letter(i+1)
                rows_limits[y_letter] = {"start": start, "stop": stop}
            return rows_limits

        def _y_max(self, x_values, x_start, x_stop):
            # For each (x) columns, find the rows corresponding to (x) axis start and stop values
            rows_limits = self._rows_limits(x_values, x_start, x_stop)

            # Create a list with all the (y) values between these rows
            y_values = []
            for col in rows_limits:
                row_min = rows_limits[col]["start"]
                row_max = rows_limits[col]["stop"]
                for cell in self.ws[col]:
                    if cell.row >= row_min and cell.row <= row_max:
                        y_values.append(cell.value)

            # Find and round the highest (y) maxima
            if y_values:
                maxima = max(y_values)
                if maxima > 200:
                    maxima = int(50 * math.ceil(maxima / 50))
                elif maxima > 50:
                    maxima = int(20 * math.ceil(maxima / 20))
                else:
                    maxima = math.ceil(maxima * 1.1)
                return maxima
            return 0


class Sheet(Workbook):
    def __init__(self, files, labels):
        super().__init__()
        self.labels = labels
        self.ws = self.active
        self.ws.title = "data"
        self._load(files)
        self._format()
        self._add_header(files)

    def _add_header(self, files):
        self.header = self._get_header(files)
        self.ws.insert_rows(1)
        for cell in self.ws["1"]:
            if cell.column > self.ws.max_column: break
            else:
                h = self.header[cell.column-1]
                cell.value = h

    def _format(self):
        for row in self.ws.iter_rows():
            for cell in row:
                value = cell.value
                try:
                    value = value.replace(",", ".")
                    value = value.replace("\ufeff", "")  # Remove BOM
                    value = float(value)
                    cell.value = value
                    cell.number_format = "0.00"
                except AttributeError:
                    pass
                except ValueError:
                    pass

    def add_stats(self, x_min, x_max):
        x_values, minimums, maximums = self.mins_maxs(1, self.ws.max_column)
        x_max = max(maximums) if not x_max else x_max
        x_min = min(minimums) if not x_min else x_min

        header = self.header[1::2]
        ws_stats = self.create_sheet("stats")
        for i, h in enumerate(header):
            ws_stats[f"A{i+1}"].value = h

        ws_stats.insert_rows(1)
        ws_stats["A1"].value = self.labels["stats title"]
        ws_stats["B1"].value = self.labels["stats x"]
        ws_stats["C1"].value = self.labels["stats y"]
        ws_stats["D1"].value = "μg/μL" ##
        for cell in ("A1", "B1", "C1", "D1"):
            ws_stats[cell].font = Font(bold=True)

        for row, column in enumerate(range(1, self.ws.max_column, 2), 2):
            x_col = get_column_letter(column)
            y_col = get_column_letter(column+1)
            values = {}
            for cell in self.ws[x_col]:
                if isinstance(cell.value, float):
                    if x_max >= cell.value >= x_min:
                        y_val = self.ws[f"{y_col}{cell.row}"].value
                        values[y_val] = cell.value
            y = max(values)
            x = values[y]
            ws_stats[f"B{row}"].value = x
            ws_stats[f"C{row}"].value = y
            ws_stats[f"D{row}"].value = f"=(C{row}-0.0456)/0.7502"  # ug/uL = (λmax – 0.0456) / 0.7502 ##
            ws_stats[f"B{row}"].number_format = "0"
            ws_stats[f"C{row}"].number_format = "0.00"
            ws_stats[f"D{row}"].number_format = "0.00"

    def mins_maxs(self, col_start, col_stop):
        # Return all the values of each (x) or (y) columns
        values = {}
        for i in range(col_start, col_stop, 2):
            letter = get_column_letter(i)
            values[letter] = []
            for cell in self.ws[letter]:
                if isinstance(cell.value, float):
                    values[letter].append(cell.value)

        minimums, maximums = [], []
        for col in values:
            minimums.append(min(values[col]))
            maximums.append(max(values[col]))
        return values, minimums, maximums


def enum(xs, start=0, step=1):
    for x in xs:
        yield (start, x)
        start += step
