#!/usr/bin/python3
import pandas
import scipy.signal

from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from ..backend import peak_base
except ValueError:
    from backend import peak_base


class RamanChart(peak_base.Chart):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.x_axis.title = self.labels["x"]
        self.y_axis.title = self.labels["y"]

    def _filter_peaks(self, values):
        average = sum(values) / len(values)
        prominence = int(average * 0.25)
        index, peaks = scipy.signal.find_peaks(values, prominence=prominence)
        return index, peaks


class RamanSheet(peak_base.Sheet):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def _get_header(self, files):
        header = []
        for f in files:
            f = Path(f)
            if f.is_file() and f.suffix.lower() == ".csv":
                header += [self.labels["x"], f.stem]
        return header

    def _load(self, files):
        combined_csv = pandas.concat([pandas.read_csv(f, sep=",", encoding="latin-1") for f in files], axis=1)
        combined_csv = combined_csv.dropna(axis='columns', how='all')  # Drop empty columns
        for row in dataframe_to_rows(combined_csv, index=False, header=False):
            self.ws.append(row)

    def add_chart(self, pos, **kwargs):
        chart = RamanChart(self, **kwargs)
        self.ws.add_chart(chart, pos)


class UVChart(peak_base.Chart):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.x_axis.title = self.labels["x"]
        self.y_axis.title = self.labels["y"]

    def _filter_peaks(self, values):
        index, peaks = scipy.signal.find_peaks(values, prominence=0.01, width=3)
        return index, peaks


class UVSheet(peak_base.Sheet):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.ws.delete_rows(2, 3)

    def _get_header(self, files=None):
        header = []
        for i in range(1, self.ws.max_column, 2):
            letter = get_column_letter(i)
            header.append(self.labels["x"])
            header.append(self.ws[f"{letter}1"].value)
        return header

    def _load(self, files):
        combined_csv = pandas.concat([pandas.read_csv(f, sep=";", encoding="latin-1") for f in files], axis=1)
        combined_csv = combined_csv.dropna(axis='columns', how='all')  # Drop empty columns
        for row in dataframe_to_rows(combined_csv, index=False):
            self.ws.append(row)

    def add_chart(self, pos, **kwargs):
        chart = UVChart(self, **kwargs)
        self.ws.add_chart(chart, pos)
