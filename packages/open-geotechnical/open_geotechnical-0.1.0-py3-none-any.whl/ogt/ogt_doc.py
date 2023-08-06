#!/usr/bin/env python3
# coding: utf-8
""""""
import os
import csv

from io import StringIO
from zipfile import ZipFile
from geojson import Feature, Point, FeatureCollection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ogt.thirdparty import bng_to_latlon

from . import utils
from .ogt_group import OGTGroup
from .ogt_cell import OGTCell
from . import ogt_validate
from . import ags4

from . import FORMATS


class OGTDocumentOptions:
    """"""
    def __init__(self):
        self.minify = False
        """Option whether to minify output Json only"""

        self.extended = True
        """Option to `extended` output with heading/group descriptions, etc"""

        self.include_stats = False
        """Stats such as groups, row count etc"""

        self.include_source = False
        """Includes `source`  and `source_cells` in output """

    def __repr__(self):
        return "<OGTDocOpts extended=%s, mini=%s>" % (self.extended, self.minify)


class OGTDocument:
    """:class:`~ogt.ogt_doc.OGTDocument`  is primary container of data (:class:`~ogt.ogt_group.OGTGroup`).



    .. code-block:: python

        from ogt import ogt_doc

        doc = ogt_doc.OGTDocument()
        err = doc.load_ags4_file("/path/to/my.ags")
        if err:
            print err
        else:
            # print the groups index
            print doc.groups_index()

            # Headings in the SAMP group
            print doc.group("SAMP").headings()

            # Return a list of units used in the document
            print doc.units()

    """

    class SORT:
        source = 0
        """Same order as source file"""

        a2z = 1
        """Alphabatical sort"""

        recommended = 2
        """Prefered sort see :func:`groups_sort`"""

    def __init__(self, sort_mode=SORT.recommended):


        self.source_refs = []
        """Source refs ef the file_name or url, nick etc """

        self.source_strings = {}
        """The original source files contents as string
            TODO: need to sort out utf8 vs ascii vs windows characters
        """

        self._groups_index_src = []
        """A list of groups in imported order"""

        self._groups_index_curr = []
        self._sort_mode = sort_mode

        self.groups_dict = {}
        """A `dict` of group code to :class:`~ogt.ogt_group.OGTGroup` instances

        :type: dict[str] => OGTGroup
        """

        self._errors_cache = None
        """A `dict` of row/col indexes with errors"""

        self.opts = OGTDocumentOptions()
        """Set default options :class:`~ogt.ogt_doc.OGTDocumentOptions` """

        self.sig_changed = None
        """A pointer to a `pyqtSignal` (only used with gui) see :meth:`emit` """

    def emit(self):
        """Emits the :attr:`sig_changed` if set"""
        print("ogtDoc.emit()", self)
        if self.sig_changed is None:
            return
        self.sig_changed.emit()

    def concat_char(self, default="+"):
        """
        :return: conantenation character (see :ref:`ags4_rule_11b`)
        :type: str
        """
        grp = self.group_from_code("TRAN")
        if len(grp.data) == 1:
            cell = grp.data[0].get("TRAN_RCON")
            if cell:
                return cell.value
        return default

    def add_get_group(self, cell_or_val, src_ref=None):
        """Adds (if not exist) or gets (if exists) a :class:`~ogt.ogt_group.OGTGroup` instance in this document

        :param grp: The group object to add
        :type grp: ~ogt.ogt_group.OGTGroup
        :return: An `Error` message is group exists, else `None`
        """

        gcell = cell_or_val if isinstance(cell_or_val, OGTCell) else OGTCell(raw=cell_or_val)

        # first check group not already there
        group_code = gcell.value

        ogtGroup = self.groups_dict.get(group_code)
        if ogtGroup is None:
            # not exist so add
            ogtGroup = OGTGroup(self, gcell, src_ref=src_ref)
            self.groups_dict[group_code] = ogtGroup

            self._groups_index_src.append(group_code)
            self._groups_index_curr.append(group_code)

        return ogtGroup

    def add_required_groups(self):
        for grp_code in GROUPS_REQUIRED:
            cell = OGTCell(raw=grp_code)
            grp = self.add_get_group(cell, src_ref="added")

            hcodes = GROUP_HEADINGS_DEFAULT.get(grp.group_code)
            cells = {}
            if hcodes:
                for hc in hcodes:
                    grpdd = ags4.DD.group(grp.group_code)
                    idx = grpdd['headings_index'].index(hc)
                    rec = grpdd['headings'][idx]
                    grp.add_heading(rec)

                    if len(grp.data) == 0:
                        cell = OGTCell(raw="")
                        cells[hc] = cell
                grp.add_data_row(cells)

        self.update_sort()

        self.emit()

    def set_sort_mode(self, sort_mode):
        """
        :param sort_mode: The index of the group
        :type sort_mode:
        Return a list of group_codes in preferred order (see :func:`~ogt.ogt_group.groups_sort`)



        """
        self._sort_mode = sort_mode
        self.update_sort()

    def update_sort(self):
        if self._sort_mode == self.SORT.a2z:
            self._groups_index_curr = sorted(self._groups_index_src)

        elif self._sort_mode == self.SORT.source:
            self._groups_index_curr = self._groups_index_src
        else:
            self._groups_index_curr = groups_sort(self._groups_index_src)

    def group_from_index(self, idx):
        """
        :param idx: The index of the group
        :type idx: int
        :rtype: OGTGroup
        :return: Groups codes index
        """
        return self.groups_dict[self._groups_index_curr[idx]]

    @property
    def groups_list(self):
        """
        :type: list[OGTGroup]
        :return: A `list` of :class:`OgtGroup` objects
        """
        return [self.groups_dict[gc] for gc in self._groups_index_curr]

    @property
    def groups_index(self):
        """
        :rtype: list[str]
        :return: Groups codes index
        """
        return self._groups_index_curr

    @property
    def groups_count(self):
        """
        :rtype: int
        :return: Returns no of groups in the document
        """
        return len(self._groups_index_src)

    def group_from_code(self, group_code):
        """
        :param group_code: Four character group code
        :type group_code: str
        :return: An instance of :class:`~ogt.ogt_group.OGTGroup` if exists, else `None`
        """
        return self.groups_dict.get(group_code)

    def proj(self):
        """Shortcut to `PROJ` group object

        :return: An instance of :class:`~ogt.ogt_group.OGTGroup` if exists, else `None`
        """
        return self.group_from_code("PROJ")

    def proj_dict(self):
        """Shortcut to `PROJ` group data

        :return: A dict with data if exists, else `None`
        """
        grpOb = self.group_from_code("PROJ")
        if not grpOb:
            return None
        if grpOb.data_rows_count > 0:  # should always be one on project row
            return grpOb.data_row_dict(0)
        return None

    def units(self):
        """Shortcut to `UNIT` group

        :rtype: tuple
        :return:
            - An instance of :class:`~ogt.ogt_group.OGTGroup` if exists, else `None`
            - `bool` = `True` if group found in document, else false
        """
        return self.group_from_code("UNIT")

    def data_types(self):
        """Shortcut to `TYPE` group

        :rtype: tuple
        :return:
            - An instance of :class:`~ogt.ogt_group.OGTGroup` if exists, else `None`
            - `bool` = `True` if group found in document, else false
        """
        return self.group_from_code("TYPE")

    def frsfabbrs(self, head_code):
        """"""
        lst = []
        grp = self.group_from_code("ABBR")
        if grp:
            for cellsDic in grp.data:
                cellH = cellsDic.get("ABBR_HDNG")
                if cellH and cellH.value == head_code:
                    dic = dict(code=cellsDic.get("ABBR_CODE").value,
                               description=cellsDic.get("ABBR_DESC").value, custom=True,
                               list="CUSTOM")
                    lst.append(dic)
        ags_abbrs = ags4.DD.picklist_vals(head_code)
        if ags_abbrs:
            lst.extend(ags_abbrs)

    def add_abbr(self, head_code, rec):
        """Add abbreviation row"""
        # TODO inset index

        existing_abbrs = self.abbrs(head_code, as_dict=True)
        if rec['code'] not in existing_abbrs:

            row = {}
            row['ABBR_HDNG'] = OGTCell(raw=head_code)
            row['ABBR_CODE'] = OGTCell(raw=rec['code'])
            row['ABBR_DESC'] = OGTCell(raw=rec['description'])
            row['ABBR_LIST'] = OGTCell(raw=rec['list'])

            grp = self.group_from_code("ABBR")

            # insert at end of heading block, otherwise end
            found_group = False
            for ridx, rec in enumerate(grp.data_rows_dict()):

                if rec['ABBR_HDNG'] == head_code:
                    found_group = True

                if rec['ABBR_HDNG'] != head_code and found_group:
                    grp.insert_data_row(ridx, row)
                    self.emit()
                    return

            grp.add_data_row(row)

    def abbrs(self, head_code, as_dict=False, code_list=False):
        """Abbtrviatiosn for heading"""
        lookup = {}
        code_idx = []

        # add abbrs from ags
        agscodes = ags4.DD.abbrs(head_code)
        if agscodes:
            for rec in agscodes:
                lookup[rec['code']] = rec
                code_idx.append(rec['code'])

        # add abbrs from this file
        grp = self.group_from_code("ABBR")
        if grp:
            for cellsDic in grp.data:
                cellH = cellsDic.get("ABBR_HDNG")
                if cellH and cellH.value == head_code:
                    code = cellsDic.get("ABBR_CODE").value
                    if code in lookup:
                        # already in ags4
                        continue
                    dic = dict(code=code, description=cellsDic.get("ABBR_DESC").value,
                               list=cellsDic.get("ABBR_LIST").value)
                    lookup[dic['code']] = dic
                    code_idx.append(dic['code'])

        if as_dict:
            return lookup
        if code_list:
            return sorted(code_idx)
        return [lookup[code] for code in sorted(code_idx)]

    # def errors(self, lidx=None, cidx=None):
    #     """"""
    #     if lidx is not None:
    #         recs = self.error_cells.get(lidx)
    #         if recs is None:
    #             return None
    #         if cidx is None:
    #             return recs
    #         return recs.get(cidx)
    #     return None

    def errors_list(self):
        """"""
        if self._errors_cache is None:
            self._update_errors_cache()
        return self._errors_cache
        lst = []
        for ridx, row in enumerate(self.cells):
            for cidx, cell in enumerate(row):
                lst.extend(cell.errors)

        return lst

        lst = []
        for lidx in sorted(self.error_cells.keys()):
            for cidx in sorted(self.error_cells[lidx].keys()):
                lst.extend(self.error_cells[lidx][cidx])
        return lst

    def _update_errors_cache(self):
        """"""
        self._errors_cache = []
        for grp in self.groups_dict.values():
            self._errors_cache.extend(grp.errors_list())

    @property
    def errors_warnings_count(self):
        """"""
        if self._errors_cache is None:
            self._update_errors_cache()

        return len(self._errors_cache)

    def clear_errors(self):
        """Clears all errors, eg before revalidation"""
        self._errors_cache = None
        for grp in self.groups_dict.values():
            grp.clear_errors()

    def validate(self):
        """WIP validate doc

        """
        # first clear all existing errors
        self.clear_errors()

        for grp in self.groups_dict.values():
            grp.validate()

        # validate each group
        for grp in self.groups_dict.values():
            grp.validate()

    def write(self, ext="json", beside=False, file_path=None,
              to_zip=False, overwrite=False):
        """Write out the data to file in the selected format

        .. Note:: **Note**

            - Either **`beside=True`** or a **`file_path`** is required, otherwise and error occurs
            - If both are provided, and error is returned

        :param ext: The file format, see :data:`~ogt.__init__.FORMATS`
        :type ext: str

        :type beside: bool
        :param beside: Save the output file alongside the original with extention appended, eg

             - Source = `/path/to/myproject.ags`
             - Output = `/path/to/myproject.ags.json`

        :param file_path: Relative or absolute path to write to including extention
        :type file_path: str

        :param include_source: If `True`, the original ags source is also included.
        :type include_source: bool

        :param to_zip: If `True`, the original and converted file are packaged in a zip
        :type to_zip: bool

        :param overwrite: If `True`, the target file is overwritten, otherwise an error is returned
        :type overwrite: bool

        :return: A tuple with

                - A `Message` string  if no errors, else `None`
                - Any `Error` that occured, otherwise `None`

        """
        self.stats()['site_geometry']

        # Do some validations
        if ext not in FORMATS:
            return None, "Error: Invalid format specified - `%s` % ext. Use %s" % ",".join(FORMATS)

        if not beside and file_path is None:
            return None, "Error: need an output, either -b or -w"

        if beside and file_path is not None:
            return None, "Error: conflict in options, either -b or -w, not BOTH"

        # make target filename's
        base_name = os.path.basename(self.source_file_path)
        target_file_path = None
        if beside:
            # File is beside the original
            if to_zip:
                target_file_path = self.source_file_path + ".zip"

            else:
                target_file_path = self.source_file_path + ".%s" % ext

        else:
            # file is from argument
            target_file_path = file_path
            base_name = os.path.basename(file_path)
            if len(base_name) == 0:
                # directory given only
                return None, "Error: Invalid file name `%s`" % target_file_path

            parts = base_name.split(".")
            if len(parts) == 1:
                # no extention
                return None, "Error: Invalid file name `%s`" % target_file_path

            # Check the extention is what we expect
            gext = parts[-1]
            if not to_zip and gext != ext:
                return None, "Error: Conflict in file name extention, expected '%s' `%s`" % (ext, target_file_path)
            elif to_zip and gext != "zip":
                # extentions mismatched eg json != yaml
                return None, "Error: Conflict in file name extention expected 'zip' `%s`" % target_file_path

        # TODO warn if not overwrite
        if not overwrite:
            if os.path.exists(target_file_path):
                return None, "Error: Target file exists - `%s` " % target_file_path

        # Excel get own function
        if ext == "xlsx":
            return self.write_excel(target_file_path)

        # convert the file to target format string blob
        blob = None
        err = None
        if ext in ["js", "json"]:
            blob, err = self.to_json()

        elif ext == "geojson":
            blob, err = self.to_geojson()

        elif ext == "yaml":
            blob, err = self.to_yaml()

        elif ext in ["ags", "ags4"]:
            blob, err = self.to_ags4()

        else:
            return None, "Error: No valid output format specified - `%s` % ext"

        if err:
            return None, err

        if to_zip:
            # create zip
            try:
                zipee = ZipFile(target_file_path, mode="w")

                # add source file
                zipee.writestr(base_name, self.source_file_path)

                # add converted file
                zipee.writestr("%s.%s" % (base_name, ext), blob)

                # write out and done
                zipee.close()

                siz = utils.file_size(target_file_path, human=True)
                return "Wrote: %s `%s`" % (siz, target_file_path), None

            except Exception as error:
                return None, "Error: %s" % str(error)
        else:
            try:
                with open(target_file_path, "w") as file:
                    file.write(blob)
                    file.close()
                siz = utils.file_size(target_file_path, human=True)
                return "Wrote: %s `%s`" % (siz, target_file_path), None

            except Exception as error:
                return None, "Error: %s" % str(error)

        return None, "Error: OOPS unexpected error"

    def to_ags4(self):
        """

        :return: The document as an ags string
        :rtype: str or None, err
        """
        out = StringIO()
        writer = csv.writer(out, delimiter=',', lineterminator='\r\n',
                            quotechar='"', quoting=csv.QUOTE_ALL)

        for group_code in self.groups_index:
            grp = self.groups_dict[group_code]

            # write "GROUP"
            writer.writerow([ags4.DESCRIPTOR.group, group_code])

            # write headings
            lst = [ags4.DESCRIPTOR.heading]
            lst.extend(grp.headings_index)
            writer.writerow(lst)

            # write units
            lst = [ags4.DESCRIPTOR.unit]
            lst.extend(grp.units_list())
            writer.writerow(lst)

            # write types
            lst = [ags4.DESCRIPTOR.data_type]
            lst.extend(grp.data_types_list())
            writer.writerow(lst)

            # write data
            for dic in grp.data:
                lst = [ags4.DESCRIPTOR.data]
                for hcode in grp.headings_index:
                    v = ""
                    cell = dic.get(hcode)
                    if cell:
                        v = cell.value
                    lst.append(v)
                writer.writerow(lst)

            writer.writerow([])

        return out.getvalue(), None

    def to_grid(self):
        grid = []
        max_col = 0

        def to_dic(lst):
            dic = {}
            for idx, v in enumerate(lst):
                dic[idx] = v
            return dic

        for group_code in self.groups_index:
            grp = self.groups_dict[group_code]

            # write "GROUP"
            grid.append(to_dic([ags4.DESCRIPTOR.group, grp.group_code_cell]))

            # write headings
            lst = [ags4.DESCRIPTOR.heading]
            for hh in grp.headings_list:
                lst.append(hh.head_cell)

            grid.append(to_dic(lst))
            if len(lst) > max_col:
                max_col = len(lst)

            # write units
            lst = [ags4.DESCRIPTOR.unit]
            for hh in grp.headings_list:
                lst.append(hh.unit_cell)
            grid.append(to_dic(lst))

            # write types
            lst = [ags4.DESCRIPTOR.data_type]
            for hh in grp.headings_list:
                lst.append(hh.data_type_cell)
            grid.append(to_dic(lst))

            # write data
            for dic in grp.data:
                lst = [ags4.DESCRIPTOR.data]
                for hcode in grp.headings_index:
                    cell = dic.get(hcode)
                    lst.append(cell)
                grid.append(to_dic(lst))

            grid.append(to_dic([]))

        return grid, max_col

    def to_dict(self):
        """
        :rtype: dict
        :return: The data in a dict, see :attr:`OGTDocument.opts` for serialisation options
        """
        # base dict to return
        dic = dict(file_name=self.source_file_path, version="ags4", groups={})

        # loop groups and add struct based on edit_mode
        for k, g in self.groups_dict.items():
            dic['groups'][k] = g.to_dict()

        # include source raw source
        if self.opts.include_source:
            pass

            # TODO
            dic['source'] = "TODO"  # self.source
            dic['source_cells'] = ""  # self.csv_rows

        # include statistics
        if self.opts.include_stats:
            dic['stats'] = self.stats()

        return dic

    def to_json(self):
        """Return the document data in :ref:`json` format

        .. code-block:: python

           ogtDoc = OGTDocument()
           ogtDoc.add_ags4_file("/my.ags")
           ogtDoc.opts.minify = False
           ogtDoc.opts.extended = False
           ags_str, err = ogtDoc.to_json()

        :rtype: (str or None, str)
        :return: A tuple with:

                - `None` if error else a `str` with :ref:`json` encoded data
                - An `error` string is error occured, else `None`
        """
        return utils.to_json(self.to_dict(), minify=self.opts.minify)

    def to_yaml(self):
        """Return the document data in :ref:`yaml` format

        :rtype: (str or None, str)
        :return: A tuple with:

                - `None` if error else a `str` with :ref:`yaml` encoded data
                - An `error` string is error occured, else `None`
        """
        return utils.to_yaml(self.to_dict())

    def get_points(self):
        """"""
        grpLoca = self.group_from_code("LOCA")
        if grpLoca is None:
            return
        lst = []

        # WSG84
        if grpLoca.has_heading("LOCAL_LAT") and grpLoca.has_heading("LOCAL_LON"):
            for rec in grpLoca.data:
                lat_s = rec.get("LOCA_LAT")
                lon_s = rec.get("LOCA_LON")
                # TODO add Point

        # BNG British National grid
        elif grpLoca.has_heading("LOCA_NATE") and grpLoca.has_heading("LOCA_NATN"):
            for rec in grpLoca.data:

                loca_id = rec.get("LOCA_ID")
                east = float(rec.get("LOCA_NATE"))
                north = float(rec.get("LOCA_NATN"))
                if east and north:
                    lat, lon = bng_to_latlon.OSGB36toWGS84(east, north)
                    lst.append(dict(lat=lat, lon=lon, east=east, north=north, loca_id=loca_id))
        return lst

    def to_geojson(self, minify=False):
        """geojson dict"""
        loca = self.group_from_code("LOCA")
        if loca is None:
            return None, "No `LOCA` Group"

        def make_feature(rec, lat, lon):
            props = dict(PointID=rec.get("LOCA_ID"), Type=rec.get("LOCA_TYPE"),
                         GroundLevel=rec.get("LOCA_GL"))
            return Feature(geometry=Point((lon, lat)), properties=props)

        features = []

        # WSG84
        if "LOCA_LAT" in loca.headings and "LOCA_LON" in loca.headings:
            for rec in loca.data:
                lat_s = rec.get("LOCA_LAT")
                lon_s = rec.get("LOCA_LON")
                if lat_s and lon_s:
                    features.append(make_feature(rec, lat_s, lon_s))

        # BNG British National grid
        elif "LOCA_NATE" in loca.headings and "LOCA_NATN" in loca.headings:
            for rec in loca.data:
                east = utils.to_int(rec.get("LOCA_NATE"))
                north = utils.to_int(rec.get("LOCA_NATN"))
                if east and north:
                    lat, lon = bng_to_latlon.OSGB36toWGS84(east, north)
                    features.append(make_feature(rec, lat, lon))

        if len(features) > 0:
            feature = FeatureCollection(features)
            return utils.to_json(feature, minify=minify)
        return None, None

    def write_excel(self, file_path):
        """Write out xlsx

        :param file_path: Full path of file to write
        :tyoe file_path: str
        :rtype: (str or None, str)
        :return: A tuple with:

                - `None` if error else a `str` with :ref:`json` encoded data
                - An `error` string is error occured, else `None`
        """
        wbook = Workbook()

        for gidx, grp_code in enumerate(self.groups_index):
            grpObj = self.groups_dict[grp_code]

            if gidx == 0:
                # By default an empty workbook has a first sheet
                sheet = wbook.active
                sheet.title = grp_code
            else:
                sheet = wbook.create_sheet(title=grp_code)

            for hidx, hd in enumerate(grpObj.headings_list()):

                # HEADINGS's
                sheet["A1"] = ags4.DESCRIPTOR.heading
                cell_ref = "%s%s" % (get_column_letter(hidx + 2), 1)
                sheet[cell_ref] = hd.head_code

                # TYPE
                sheet["A2"] = ags4.DESCRIPTOR.data_type
                cell_ref = "%s%s" % (get_column_letter(hidx + 2), 2)
                sheet[cell_ref] = hd.data_type

                # UNIT
                sheet["A3"] = ags4.DESCRIPTOR.unit
                cell_ref = "%s%s" % (get_column_letter(hidx + 2), 3)
                sheet[cell_ref] = hd.unit

                # DATA
                didx = 4
                for drow in grpObj.data_rows_dict():
                    sheet["A%s" % didx] = ags4.DESCRIPTOR.data
                    cell_ref = "%s%s" % (get_column_letter(hidx + 2), didx)
                    sheet[cell_ref] = drow[hd.head_code]

                    didx += 1
        try:
            wbook.save(file_path)
            return "Wrote `%s`" % file_path, None
        except Exception as error:
            return "Error: %s writing excel `%s`" % (error, file_path)

    def stats(self):
        """Statistics TODO"""
        dic = {}

        # Number of locations
        locaGrp = self.groups_dict.get("LOCA")
        if locaGrp is None:
            dic['locations'] = None
        else:
            recs = locaGrp.data_column_from_head_code("LOCA_ID")
            dic['locations'] = dict(count=len(recs), data=recs)

        # Data rows
        lst = []
        for group_key in sorted(self.groups_dict.keys()):
            grp = self.groups_dict.get(group_key)
            lst.append(dict(GROUP=group_key, count=len(grp.data)))
        dic['data'] = lst

        # Sample Types
        grp = self.groups_dict.get("SAMP")
        if not grp:
            dic['sample_types'] = None
        else:
            d = {}
            reccells = grp.data_column_from_head_code("SAMP_TYPE")
            recs = [cell.value for cell in reccells]

            for st in sorted(recs):
                if st not in d:
                    d[st] = 0
                d[st] += 1
            dic['sample_types'] = d

        # Site Geom
        d = {}
        # TODO X.Y.Z
        d['LOCA_LOCX'] = "todo"
        d['LOCA_LOCY'] = "todo"
        d['LOCA_LOCZ'] = "todo"

        # National Grid
        def calc_ng_stats(recs):
            # TODO - need to check type casting ?
            if recs is None:
                return None

            ds = {}
            ds['min'] = None if len(recs) == 0 else min(recs)
            ds['max'] = None if len(recs) == 0 else max(recs)
            ds['row_count'] = len(recs)
            ds['rows_with_data'] = 0
            ds['rows_without_data'] = 0
            for rec in recs:
                if not rec:
                    ds['rows_without_data'] += 1
                else:
                    ds['rows_with_data'] += 1
            return ds

        dic['site_geometry'] = None
        if locaGrp:
            recs = locaGrp.data_column_from_head_code("LOCA_NATE", as_cells=False)
            d['LOCA_NATE'] = calc_ng_stats(recs)

            recs = locaGrp.data_column_from_head_code("LOCA_NATN", as_cells=False)
            d['LOCA_NATN'] = calc_ng_stats(recs)

            recs = locaGrp.data_column_from_head_code("LOCA_GL", as_cells=False)
            d['LOCA_GL'] = calc_ng_stats(recs)

            dic['site_geometry'] = d

        # GEOL
        grp = self.groups_dict.get("GEOL")
        if not grp:
            dic['geol'] = None
        else:
            recs = grp.data_column("LOCA_ID")
            locs = dic['locations']['data']
            ll = []
            for l in locs:
                if l not in recs:
                    if l not in ll:
                        ll.append(l)
            dic['geol'] = dict(no_entries=ll if len(ll) > 0 else None)

        # SAMP
        grp = self.groups_dict.get("SAMP")
        if not grp:
            dic['samp'] = None
        else:
            recs = grp.data_column_from_head_code("LOCA_ID")
            locs = dic['locations']['data']
            ll = []
            for l in locs:
                if l not in recs:
                    if l not in ll:
                        ll.append(l)
            dic['samp'] = dict(no_entries=ll if len(ll) > 0 else None)

        # Unused Groups
        all_g = ags4.DD.groups_dict()
        dic['unused_groups'] = None
        ags_groups = all_g.keys()
        dic['unused_groups'] = sorted(list(set(ags_groups) - set(self.groups_dict.keys())))

        return dic

    def add_ags4_file(self, ags4_file_path):
        """Loads document from an :term:`ags4` formatted file

        :param ags4_file_path: absolute or relative path to file, will be at source_file_path
        :type ags4_file_path: str
        :rtype: str or None
        :return: A str if an error else None


        .. todo:: deal with utf-8 and windows.chars as ags is ascii only
                  another BIG issue is adding multiple files.. + source...s
        """
        # This function read's the file and call `add_ags4_string`

        try:
            with open(ags4_file_path, "r") as file:
                contents = file.read()
                self.add_ags4_string(contents, ags4_file_path)
                return None

        except IOError as error:
            return "%s - `%s`" % (str(error), ags4_file_path)

        # should never happen
        return "WTF in `add_ags4_file`"

    @property
    def source_file_path(self):
        """TODO wourkaround atmo"""
        return self.source_refs[0]

    def add_ags4_string(self, ags4_str, src_ref):
        """Load  document from an :term:`ags4` formatted string

        :param ags4_str: ags4 formatted string to load
        :type ags4_str: str
        :param src_ref: filename, eg excel, ags etc TODO
        :type src_ref: str
        :rtype: str or None
        :return: Any error  else `None`
        """

        self.source_refs.append(src_ref)
        self.source_strings[src_ref] = ags4_str

        # Note: the csv parser cant deal columns of different length,
        # Step 1:
        #  - split ags_string into lines
        #  - and parse `each line` into csv
        xcells = []
        for lidx, line in enumerate(ags4_str.split("\n")):

            # removing and trailing whitespace eg \r
            # were in nix land, so reassemble with CRLF when dumping to ags
            stripped = line.strip()

            if stripped == "":
                continue

            # decode csv line
            reader = csv.reader(StringIO(stripped))
            for row in reader:
                csv_row = row
                break

            # walk row and make OgtCell outta each bit of stuff
            row_cells = []
            for cidx, val in enumerate(csv_row):
                row_cells.append(OGTCell(src_ref=src_ref, lidx=lidx, cidx=cidx, raw=val))
            xcells.append(row_cells)
        row_cells = None

        # --------------------
        # next walk and clean

        def row_2_upper(rcells):
            for cell in rcells:
                ogt_validate.to_upper(cell)

        loop_group = None
        for lidx, row_cells in enumerate(xcells):
            lenny = len(row_cells)
            if lenny == 0:
                continue

            # validate its a descriptor
            dcell = row_cells[0]
            valid = True
            if not valid:
                continue
            else:
                ogt_validate.to_upper(dcell)
                descriptor = dcell.value

                if descriptor == ags4.DESCRIPTOR.group:
                    row_2_upper([row_cells[1]])
                    loop_group = self.add_get_group(row_cells[1], src_ref=src_ref)

                elif descriptor == ags4.DESCRIPTOR.heading:
                    row_2_upper(row_cells[1:])
                    head_list = loop_group.add_get_headings_row(row_cells[1:], src_ref=src_ref)

                elif descriptor == ags4.DESCRIPTOR.unit:
                    for cidx, lcell in enumerate(row_cells[1:]):
                        lcell.parentHeading = loop_group.headings_dict[head_list[cidx]]
                        loop_group.headings_dict[head_list[cidx]].set_unit(lcell)

                elif descriptor == ags4.DESCRIPTOR.data_type:
                    row_2_upper(row_cells[1:])
                    for cidx, lcell in enumerate(row_cells[1:]):
                        lcell.parentHeading = loop_group.headings_dict[head_list[cidx]]
                        loop_group.headings_dict[head_list[cidx]].set_data_type(lcell)

                elif descriptor == ags4.DESCRIPTOR.data:
                    datarow = {}
                    for cidx, lcell in enumerate(row_cells[1:]):
                        lcell.parentHeading = loop_group.headings_dict[head_list[cidx]]
                        datarow[head_list[cidx]] = lcell
                    loop_group.data.append(datarow)

                else:
                    print("LOST row", row_cells)
                    # loop_group.add_lost_row(rrow, lidx)

        self.set_sort_mode(self.SORT.recommended)
        self.generate_samp_ids()

        return None

    def generate_samp_ids(self):
        """"""
        # attempt to SAMP_ID
        # grp = self.group("SAMP")
        for gcode, grp in self.groups_dict.items():
            # if grp:
            for headObj in grp.headings_dict.values():
                if headObj.head_code == "SAMP_ID":
                    for ridx in range(0, grp.data_rows_count):
                        cells_dic = grp.data[ridx]
                        id_cell = cells_dic.get("SAMP_ID")

                        if id_cell is None:
                            # panix
                            continue
                        delim = "~"
                        if not id_cell.raw:
                            s = "@"
                            cell = cells_dic.get("LOCA_ID")
                            s += "?loca_id?" if cell is None else cell.value
                            s += delim

                            cell = cells_dic.get("SAMP_TYPE")
                            s += "?type?" if cell is None else cell.value
                            s += delim

                            cell = cells_dic.get("SAMP_REF")
                            if cell and cell.value:
                                s += "ref:" + cell.value
                            else:
                                cell = cells_dic.get("SAMP_TOP")
                                if cell:
                                    s += "top:" + cell.value
                                else:
                                    s += "ridx: %s" % ridx

                            id_cell.value = str(s)


def create_doc_from_ags4_file(ags_file_path):
    """Convenience function to create and load an OGTDocument from an ags file

    .. code-block:: python

        doc, err = ogt_doc.create_doc_from_ags4_file("/path/to/my.ags")
        if err:
            print err
        else:
            print doc.group("PROJ")
    """
    doc = OGTDocument()
    err = doc.add_ags4_file(ags_file_path)
    return doc, err


def create_doc_from_ags4_string(ags4_string, file_name=None):
    """Convenience function to create and load an OGTDocument from an ags string

    .. code-block:: python

        doc, err = ogt_doc.create_doc_from_ags4_string(ags4_string)
        if err:
            print err
        else:
            print doc.group("PROJ")
    """
    doc = OGTDocument()
    doc.add_ags4_string(ags4_string, file_name)
    return doc, None


def create_doc_from_json_file(json_file_path):
    """Creates a document from a :ref:`json` formatted file

    .. code-block:: python

        doc, err = ogt_doc.create_doc_from_json_file("/path/to/my.json")
        if err:
            print err

    :param json_file_path: absolute or relative path to file
    :type json_file_path: str
    :rtype: tuple
    :return: A `tuple` containing

        - An :class:`~ogt.ogt_doc.OGTDocument` object on success, else `None`
        - An `Error` message if error, otherwise `None`
    """
    data, err = utils.read_json_file(json_file_path)
    if err:
        return None, err

    groups = data.get('groups')
    if groups is None:
        return None, "Error: no `groups` key in json file"

    doc = OGTDocument()


    for group_code in groups.keys():

        group = groups[group_code]

        grp = OGTGroup(group_code)
        # TODO doc.append_group(grp)

        # add units + also headings
        for head_code in group['UNIT'].keys():
            valu = str(group['UNIT'][head_code])
            # TODO grp.units[head_code] = valu
            # TODO grp.headings[head_code] = valu

        # add TYPE
        for head_code in group['TYPE'].keys():
            valu = str(group['TYPE'][head_code])
            # TODO grp.data_types_dict[head_code] = valu

        # add data
        for rec in group['DATA']:
            dic = {}
            for head_code in rec.keys():

                dic[head_code] = str(rec[head_code])

            grp.data.append(dic)

    return doc, None


GROUPS_FIRST_CODES = ["PROJ", "TRAN", "LOCA", "HDPH", "GEOL", "SAMP"]
"""The groups to appear first, see :func:`groups_sort` """

GROUPS_LAST_CODES = ["ABBR", "DICT", "FILE", "TYPE", "UNIT"]
"""The groups to appear last, see :func:`groups_sort`"""

GROUPS_REQUIRED = ["PROJ", "TRAN", "TYPE", "UNIT", "ABBR"]
"""Mandatory groups"""

GROUP_HEADINGS_DEFAULT = {
    "PROJ": ["PROJ_ID", "PROJ_NAME"],
    "TRAN": ["TRAN_DATE", "TRAN_STAT"],
    "TYPE": ["TYPE_TYPE", "TYPE_DESC"],
    "UNIT": ["UNIT_UNIT", "UNIT_DESC"],
    "ABBR": ["ABBR_HDNG",  "ABBR_CODE", "ABBR_DESC", "ABBR_LIST"]
}
"""Default headings to add to :attr:`GROUPS_REQUIRED`"""


def groups_sort(unsorted_groups):
    """Returns a 'preferred' order for the group keys

    .. note::

        This is a bit hacky as the ags spec does not define
        any group order, however PROJ does sensibly tend to come first etc.

        So this return a list with

        - at the start - :attr:`GROUPS_FIRST_CODES`
        - at end - :attr:`GROUPS_LAST_CODES`
        - and everything else in the middle sorted alphabetically
        - One idea might be to push insitu stuff to front, schedule
          and lab stuff next, monitoring stuff to end to reflect project progression

    :return: group codes in preffered order
    :rtype: list[str]
    """

    # make a newlist/copy of the groups in doc
    grps = unsorted_groups[:]

    # the start groups
    ret_start = []
    for group_code in GROUPS_FIRST_CODES:
        if group_code in grps:
            # exists so add to start, and remove from list
            ret_start.append(group_code)
            grps.remove(group_code)

    # the end groups
    ret_end = []
    for group_code in GROUPS_LAST_CODES:
        if group_code in grps:
            # exists so add to end, and remove from list
            ret_end.append(group_code)
            grps.remove(group_code)

    return ret_start + sorted(grps) + ret_end
