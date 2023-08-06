# -*- coding: utf-8 -*-


import os
import datetime

from Qt import QtGui, QtWidgets, QtCore, Qt, pyqtSignal

import openpyxl  # import load_workbook
import openpyxl.utils

from ogt import ags4
import xwidgets
import ogtgui_widgets






class ExcelWorkbookWidget(QtWidgets.QWidget):

    def __init__(self, parent=None, filename=None):
        QtWidgets.QWidget.__init__(self, parent)

        self.debug = False
        self.filename = filename
        self.setObjectName("ExcelWorkbookWidget")

        self.mainLayout = xwidgets.vlayout()
        self.setLayout(self.mainLayout)


        self.tabWidget = QtWidgets.QTabWidget()
        self.mainLayout.addWidget(self.tabWidget)



    def load_workbook(self, filename=None):


        progress = QtWidgets.QProgressDialog(self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setLabelText("Loading...\n%s" % filename)
        progress.setRange(0, 0)
        progress.setValue(0)
        progress.forceShow()

        #filename = str(filename)
        #print "==", self.filename, type(self.filename), self
        wb = openpyxl.load_workbook(filename=self.filename, data_only=True)
        for sheet in wb.worksheets:
            sheetWidget = ExcelSheetWidget(sheet=sheet)
            self.tabWidget.addTab(sheetWidget, sheet.title)

        progress.hide()



class ExcelSheetWidget(QtWidgets.QWidget):

    def __init__(self, parent=None, sheet=None):
        QtWidgets.QWidget.__init__(self, parent)


        self.setObjectName("ExcelSheetWidget")

        self.sheet = sheet

        self.mainLayout = xwidgets.vlayout()
        self.setLayout(self.mainLayout)

        self.table = QtWidgets.QTableWidget()
        self.mainLayout.addWidget(self.table)

        self.load_sheet()

    def load_sheet(self):
        #print "load_sheet", self.sheet, self
        #print self.sheet['A1'].value, self.sheet['A2']

        last_col = None

        maxi = 100
        for ridx in range(0, maxi):
            if self.table.rowCount() < ridx + 1:
                self.table.setRowCount(ridx + 1)
            for cidx in range(0, maxi):
                if self.table.columnCount() < cidx + 1:
                   self.table.setColumnCount(cidx + 1)
                   item = QtWidgets.QTableWidgetItem()
                   item.setText(openpyxl.utils.get_column_letter(cidx + 1))
                   self.table.setHorizontalHeaderItem(cidx, item)
                #cell_ref = ogt_excel.rowcol_to_cell(ridx, cidx)
                cell_ref = openpyxl.utils.get_column_letter(cidx + 1) + str(ridx + 1)
                #print("cell=", cell_ref)
                v = self.sheet[cell_ref].value
                #print ridx, cidx, cell_ref, v
                if v != None and isinstance(v, str) and v.upper() == "STOP":
                    last_col = cidx
                    #dsa
                else:
                    if v == None:
                        cv = ""
                    #elif isinstance(v, long):
                    #    cv = str(v)

                    elif isinstance(v, float):
                        cv = str(v)
                    elif isinstance(v, datetime.datetime):
                        cv = str(v)
                    else:
                        cv = v
                    item = QtWidgets.QTableWidgetItem()
                    item.setText( str(cv))
                    self.table.setItem(ridx, cidx, item)
                    #print(item, cv)

                    if ridx == 0 and v != None:
                        #print("v=", v)
                        if v.upper() == v:
                            grp = ags4.DD.heading(v)
                            #print("grp=", grp)
                            if grp:
                                #
                                #print(YES)
                                item.setBackground(QtGui.QColor("#C4FAA9"))



class ExcelBrowserWidget( QtWidgets.QWidget ):
    """T"""

    def __init__( self, parent=None, file_path=None, path_browser=False):
        super().__init__( parent )

        self.debug = False
        self.setObjectName("ExcelBrowserWidget")

        self.mainLayout = xwidgets.vlayout()
        self.setLayout(self.mainLayout)

        self.splitter = QtWidgets.QSplitter()
        self.mainLayout.addWidget(self.splitter, 10)

        self.tabWidget = QtWidgets.QTabWidget()
        self.splitter.addWidget(self.tabWidget)

        if path_browser:
            self.pathBrowse = ogtgui_widgets.ExpPathBrowseWidget()
            self.splitter.addWidget(self.pathBrowse)
            self.pathBrowse.sigOpenFile.connect(self.on_open_excel)

        self.splitter.setStretchFactor(0, 4)
        if path_browser:
            self.splitter.setStretchFactor(1, 1)




        if file_path:
            self.on_open_excel(file_path)

    def init_load(self):
        pass


    def on_open_excel(self, file_path):
        """
        open excel file
        """
        filename = str(file_path)

        widget = ExcelWorkbookWidget(filename=filename)
        self.tabWidget.addTab(widget, os.path.basename(filename))
        widget.load_workbook()
        #progress.hide()

class ExcelBrowserDialog( QtWidgets.QDialog ):
    """T"""

    def __init__( self, parent=None, file_path=None):
        super().__init__( parent )

        #self.file_path = file_path
        assert file_path != None

        self.mainLayout = xwidgets.vlayout()
        self.setLayout(self.mainLayout)

        self.excelBrowser = ExcelBrowserWidget(self, file_path=file_path)
        self.mainLayout.addWidget(self.excelBrowser)


        self.setWindowState(Qt.WindowMaximized)
